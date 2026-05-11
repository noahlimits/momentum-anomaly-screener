from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import AppConfig
from .database import Database
from .recommendations import StrategyResult
from .utils import report_filename, safe_filename


ACTION_COLORS = {
    "EXIT": "FFC7CE",
    "HOLD": "E2F0D9",
    "BUY": "C6EFCE",
    "ADD": "C6EFCE",
    "RESIZE_UP": "FFEB9C",
    "RESIZE_DOWN": "FFEB9C",
    "BLOCKED_BY_REGIME": "D9E1F2",
    "DATA_ERROR": "D9D2E9",
    "NO_CASH": "F2F2F2",
    "SKIP_NO_CASH": "F2F2F2",
    "WATCHLIST": "DDEBF7",
}


def write_workbook(result: StrategyResult, db: Database, config: AppConfig, output_path: Path | None = None) -> Path:
    path = output_path or _default_workbook_path(result, db, config)
    path.parent.mkdir(parents=True, exist_ok=True)

    dashboard = _dashboard_frame(result)
    portfolio_check = _portfolio_check_frame(result)
    side_by_side = side_by_side_frame(result)
    trade_suggestions = _recommendations_frame(result)
    rankings = _rankings_frame(result)
    candidates = _candidates_frame(result)
    universe = _universe_frame(result)
    settings = pd.DataFrame(sorted(result.settings.items()), columns=["setting", "value"])
    run_log = _run_log_frame(db, result.portfolio_id)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        dashboard.to_excel(writer, sheet_name="Dashboard", index=False)
        portfolio_check.to_excel(writer, sheet_name="Portfolio_Check", index=False)
        side_by_side.to_excel(writer, sheet_name="Current_vs_Target", index=False)
        trade_suggestions.to_excel(writer, sheet_name="Trade_Suggestions", index=False)
        rankings.to_excel(writer, sheet_name="Rankings", index=False)
        candidates.to_excel(writer, sheet_name="Candidates", index=False)
        universe.to_excel(writer, sheet_name="Universe", index=False)
        settings.to_excel(writer, sheet_name="Settings", index=False)
        run_log.to_excel(writer, sheet_name="Run_Log", index=False)
        _format_workbook(writer.book)

    if result.portfolio_id is not None:
        db.update_portfolio_review(result.portfolio_id, result.portfolio_value, str(path))
    return path


def _dashboard_frame(result: StrategyResult) -> pd.DataFrame:
    exits = sum(1 for item in result.recommendations if item.action == "EXIT")
    additions = sum(1 for item in result.recommendations if item.action in {"BUY", "ADD"})
    resize = sum(1 for item in result.recommendations if item.action in {"RESIZE_UP", "RESIZE_DOWN"})
    data_errors = sum(1 for item in result.recommendations if item.action == "DATA_ERROR") + sum(1 for item in result.scores if item.data_status != "OK")
    estimated_cash = result.portfolio_value - sum(item.target_value for item in result.recommendations if item.action not in {"EXIT", "DATA_ERROR"})
    rows = [
        ("run_id", result.run_id),
        ("portfolio_id", result.portfolio_id),
        ("mode", result.mode),
        ("selected_universe", result.universe_profile["display_name"]),
        ("regime_proxy", result.regime.proxy),
        ("regime_proxy_close", result.regime.proxy_close),
        ("regime_proxy_ma", result.regime.proxy_ma),
        ("regime_status", result.regime.status),
        ("new_buys_allowed", result.regime.allows_new_buys),
        ("portfolio_value", result.portfolio_value),
        ("calculated_portfolio_value", result.calculated_portfolio_value),
        ("cash_adjustment", result.cash_adjustment),
        ("mirror_positions_count", len(result.holdings)),
        ("exits_flagged", exits),
        ("additions_suggested", additions),
        ("resize_flags", resize),
        ("estimated_cash", estimated_cash),
        ("data_errors", data_errors),
    ]
    return pd.DataFrame(rows, columns=["metric", "value"])


def _portfolio_check_frame(result: StrategyResult) -> pd.DataFrame:
    score_by_ticker = {score.ticker: score for score in result.scores}
    rows: list[dict[str, Any]] = []
    for rec in result.recommendations:
        score = score_by_ticker.get(rec.ticker)
        if rec.action in {"BUY", "ADD", "NO_CASH", "SKIP_NO_CASH", "WATCHLIST", "BLOCKED_BY_REGIME"} and rec.current_shares == 0:
            continue
        rows.append(_combined_row(rec, score))
    return pd.DataFrame(rows)


def _recommendations_frame(result: StrategyResult) -> pd.DataFrame:
    rows = [
        {
            "action": rec.action,
            "ticker": rec.ticker,
            "current_shares": rec.current_shares,
            "target_shares": rec.target_shares,
            "share_change": rec.share_change,
            "current_price": rec.current_price,
            "approximate_dollar_value": rec.target_value,
            "target_weight": rec.target_weight,
            "reason": rec.reason,
        }
        for rec in result.recommendations
    ]
    return pd.DataFrame(rows)


def side_by_side_frame(result: StrategyResult) -> pd.DataFrame:
    score_by_ticker = {score.ticker: score for score in result.scores}
    current_by_ticker = {holding["ticker"]: holding for holding in result.holdings}
    rec_by_ticker = {rec.ticker: rec for rec in result.recommendations}
    tickers = sorted(set(current_by_ticker) | set(rec_by_ticker))
    rows = []
    for ticker in tickers:
        holding = current_by_ticker.get(ticker)
        rec = rec_by_ticker.get(ticker)
        score = score_by_ticker.get(ticker)
        current_shares = float(holding["shares"]) if holding else 0.0
        target_shares = rec.target_shares if rec else current_shares
        price = rec.current_price if rec else (score.price if score else None)
        rows.append(
            {
                "ticker": ticker,
                "current_shares": current_shares,
                "current_price": price,
                "current_value": current_shares * price if price is not None else None,
                "target_shares": target_shares,
                "target_price": price,
                "target_value": target_shares * price if price is not None else None,
                "share_change": target_shares - current_shares,
                "action": rec.action if rec else "HOLD",
                "reason": rec.reason if rec else "",
                "rank": score.rank if score else None,
                "above_100dma": score.above_100dma if score else None,
                "top_20pct": score.top_20pct if score else None,
                "gap_pass": score.gap_pass if score else None,
            }
        )
    return pd.DataFrame(rows)


def _rankings_frame(result: StrategyResult) -> pd.DataFrame:
    rows = [
        {
            "rank": score.rank,
            "ticker": score.ticker,
            "company": score.company_name,
            "sector": score.sector,
            "price": score.price,
            "momentum_score": score.momentum_score,
            "annualized_slope": score.annualized_slope,
            "r_squared": score.r_squared,
            "atr20": score.atr20,
            "ma100": score.ma100,
            "above_100dma": score.above_100dma,
            "max_single_day_move": score.gap_max_abs_move,
            "gap_pass": score.gap_pass,
            "top_20pct": score.top_20pct,
            "eligible": score.eligible,
            "data_status": score.data_status,
        }
        for score in result.scores
    ]
    return pd.DataFrame(rows)


def _candidates_frame(result: StrategyResult) -> pd.DataFrame:
    score_by_ticker = {score.ticker: score for score in result.scores}
    rows = []
    for rec in result.recommendations:
        if rec.action not in {"BUY", "ADD", "WATCHLIST", "SKIP_NO_CASH", "NO_CASH", "BLOCKED_BY_REGIME"}:
            continue
        score = score_by_ticker.get(rec.ticker)
        rows.append(
            {
                "rank": score.rank if score else None,
                "ticker": rec.ticker,
                "price": rec.current_price,
                "atr20": score.atr20 if score else None,
                "target_shares": rec.target_shares,
                "target_value": rec.target_value,
                "target_weight": rec.target_weight,
                "momentum_score": score.momentum_score if score else None,
                "candidate_status": rec.action,
                "reason": rec.reason,
            }
        )
    return pd.DataFrame(rows)


def _universe_frame(result: StrategyResult) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "ticker": item.ticker,
                "company_name": item.company_name,
                "sector": item.sector,
                "source": item.source,
                "source_date": item.source_date,
                "active": item.active,
            }
            for item in result.constituents
        ]
    )


def _run_log_frame(db: Database, portfolio_id: int | None = None) -> pd.DataFrame:
    where = ""
    params: tuple[Any, ...] = ()
    if portfolio_id is not None:
        where = "WHERE portfolio_id = ?"
        params = (portfolio_id,)
    with db.connect() as conn:
        rows = conn.execute(
            f"""
            SELECT run_datetime, universe_id, portfolio_value, regime_status,
                   holdings_reviewed, exits_flagged, additions_suggested,
                   data_errors, cash_adjustment, notes
            FROM run_log
            {where}
            ORDER BY run_id DESC
            LIMIT 50
            """,
            params,
        ).fetchall()
    return pd.DataFrame([dict(row) for row in rows])


def _default_workbook_path(result: StrategyResult, db: Database, config: AppConfig) -> Path:
    if result.portfolio_id is None:
        return report_filename(result.universe_id, result.mode, config.reports_dir)
    portfolio = db.portfolio(result.portfolio_id)
    filename = f"portfolio_{result.portfolio_id}_{safe_filename(portfolio['name'])}.xlsx"
    return config.reports_dir / filename


def _combined_row(rec: Any, score: Any) -> dict[str, Any]:
    return {
        "ticker": rec.ticker,
        "shares": rec.current_shares,
        "current_price": rec.current_price,
        "current_value": rec.current_shares * rec.current_price if rec.current_price else None,
        "rank": score.rank if score else None,
        "percentile_rank": score.percentile_rank if score else None,
        "momentum_score": score.momentum_score if score else None,
        "above_100dma": score.above_100dma if score else False,
        "gap_pass": score.gap_pass if score else False,
        "top_20pct": score.top_20pct if score else False,
        "in_universe": score.in_universe if score else False,
        "current_action": rec.action,
        "reason": rec.reason,
        "target_shares": rec.target_shares,
        "share_change": rec.share_change,
        "target_value": rec.target_value,
        "target_weight": rec.target_weight,
    }


def _format_workbook(workbook: Any) -> None:
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAD3")
        for column_cells in worksheet.columns:
            values = [str(cell.value) for cell in column_cells if cell.value is not None]
            width = min(45, max([10, *[len(value) + 2 for value in values]]))
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        _format_numbers(worksheet)
        _apply_action_colors(worksheet)


def _format_numbers(worksheet: Any) -> None:
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            header = worksheet.cell(row=1, column=cell.column).value
            if header in {"target_weight", "percentile_rank", "annualized_slope", "r_squared", "momentum_score", "max_single_day_move"}:
                cell.number_format = "0.00%"
            elif header in {"price", "current_price", "current_value", "target_value", "approximate_dollar_value", "portfolio_value", "estimated_cash", "regime_proxy_close", "regime_proxy_ma", "atr20", "ma100"}:
                cell.number_format = "$#,##0.00"
            elif header and "shares" in str(header):
                cell.number_format = "0.00"


def _apply_action_colors(worksheet: Any) -> None:
    headers = [cell.value for cell in worksheet[1]]
    action_headers = {"action", "current_action", "candidate_status"}
    action_col = next((index + 1 for index, header in enumerate(headers) if header in action_headers), None)
    if not action_col or worksheet.max_row < 2:
        return
    action_letter = get_column_letter(action_col)
    range_ref = f"A2:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    for action, color in ACTION_COLORS.items():
        formula = f'${action_letter}2="{action}"'
        worksheet.conditional_formatting.add(
            range_ref,
            FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=color)),
        )
